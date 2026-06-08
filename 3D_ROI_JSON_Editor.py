from PyQt6.QtWidgets import (
    QApplication, QWidget, QPushButton, QLabel,
    QVBoxLayout, QHBoxLayout, QFileDialog, QLineEdit,
    QFormLayout, QScrollArea, QGroupBox, QMessageBox,
    QDialog, QDialogButtonBox, QComboBox, QSizePolicy
)
from PyQt6.QtCore import Qt, QTimer, pyqtSignal
from PyQt6.QtGui import QFont, QPainter, QColor, QBrush
import json
import sys
import re
import openpyxl

class ToggleSwitch(QWidget):
    toggled = pyqtSignal(bool)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedSize(60, 30)
        self.isOn = False
        self.enabled = False

        # Renk tanımlamaları
        self.ACTIVE_COLOR = "#5E81AC"
        self.INACTIVE_COLOR = "#4C566A" 
        self.HANDLE_COLOR_ENABLED = "#ECEFF4"
        self.HANDLE_COLOR_DISABLED = "#D8DEE9"
        
    def setEnabled(self, enabled):
        self.enabled = enabled
        self.update()
        
    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        
        # Arka plan rengi
        if not self.enabled:
            color = self.INACTIVE_COLOR
        elif self.isOn:
            color = self.ACTIVE_COLOR
        else:
            color = self.INACTIVE_COLOR
            
        # Arka planı çiz
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QBrush(QColor(color)))
        painter.drawRoundedRect(0, 0, self.width(), self.height(), 15, 15)
        
        # Toggle düğmesi rengi
        handle_color = self.HANDLE_COLOR_ENABLED if self.enabled else self.HANDLE_COLOR_DISABLED
        painter.setBrush(QBrush(QColor(handle_color)))
        
        # Toggle düğmesinin konumu
        x_pos = self.width()-28 if self.isOn else 2
        painter.drawEllipse(x_pos, 2, 26, 26)
            
    def mousePressEvent(self, event):
        if self.enabled:
            self.isOn = not self.isOn
            self.toggled.emit(self.isOn)
            self.update()


class RoiEditor(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("3D ROI Editor")
        self.file_path = None
        self.json_data = {}
        self.excel_data = []
        self.excel_headers = []
        self.auto_import_enabled = False
        self.auto_import_enabled_plc = False
        self.auto_import_timer = QTimer(self)
        self.auto_import_timer.timeout.connect(self.auto_import_excel)
        
        # PLC için ayrı timer
        self.auto_import_timer_plc = QTimer(self)
        self.auto_import_timer_plc.timeout.connect(self.auto_import_excel_plc)

        self.input_fields = {
            "roi3d_in_camera_coord": {"half_lengths": [], "roi_center_pose": []},
            "roi3d_in_robot_coord": {"half_lengths": [], "roi_center_pose": []},
            "piece_info": {"part_code": None, "length": None, "width": None, "height": None}
        }
        self.init_ui()

        # Timer oluşturma ve bağlama
        self.roi_update_timer = QTimer(self)
        self.roi_update_timer.timeout.connect(self.update_roi_center_pose)
        self.roi_update_timer.start(500)  # Yarım saniyede bir çağır

    def init_ui(self):
        font = QFont("Segoe UI", 10)
        self.setStyleSheet("""
            QWidget {
                background-color: #2E3440;
                color: #ECEFF4;
            }
            QPushButton {
                background-color: #5E81AC;
                color: white;
                border: none;
                padding: 20px 20px;
                border-radius: 5px;
                font:17px;
                font-style:bold;
            }
            QPushButton:hover {
                background-color: #81A1C1;
            }
            QLineEdit {
                background-color: #3B4252;
                color: #E5E9F0;
                border: 1px solid #4C566A;
                border-radius: 3px;
                padding: 5px;
            }
            QLabel {
                font-size: 14px;
                color: #D8DEE9;
            }
            QGroupBox {
                border: 1px solid #4C566A;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 3px;
            }
            QLineEdit[error="true"] {
                border: 2px solid #BF616A;
                background-color: #4C566A;
            }
        """)

        top_right_layout = QVBoxLayout()
        # header_layout'a ekle
        header_layout = QVBoxLayout()
        header_layout.addLayout(top_right_layout)

        self.lbl_file = QLabel("No JSON File Loaded...")
        self.lbl_file.setFont(font)
        header_layout.addWidget(self.lbl_file)

        button_layout = QHBoxLayout()
        # ...butonlarınızı ekleyin...
        header_layout.addLayout(button_layout)

        button_layout = QHBoxLayout()
        self.btn_open = QPushButton("Load JSON File")
        self.btn_open.setFixedSize(220, 80)
        self.btn_open.clicked.connect(self.open_json_file)

        self.btn_open_excel = QPushButton("Load Excel File(PART)")
        self.btn_open_excel.setFixedSize(220, 80)
        self.btn_open_excel.clicked.connect(self.load_excel_file)
        self.btn_open_excel.setEnabled(False)
        self.btn_open_excel.setToolTip("To load the Excel file, first open the JSON file.")

        # Yeni "Load Excel File" butonu
        self.btn_open_excel_2 = QPushButton("Load Excel File(PLC)")
        self.btn_open_excel_2.setFixedSize(220, 80)
        self.btn_open_excel_2.clicked.connect(self.load_excel_file_plc)  # Aynı fonksiyonu kullan
        self.btn_open_excel_2.setEnabled(False)
        self.btn_open_excel_2.setToolTip("To load the Excel file, first open the JSON file.")

        self.combo_excel_data = QComboBox()
        self.combo_excel_data.setPlaceholderText("Loadign Excel Data...")
        self.combo_excel_data.currentIndexChanged.connect(self.on_excel_selection_changed)

        self.btn_save = QPushButton("Save Changes")
        self.btn_save.setFixedSize(220, 80)
        self.btn_save.clicked.connect(self.save_changes)
        self.btn_save.setEnabled(False)

        self.btn_next = QPushButton("OFFSET")
        self.btn_next.setFixedSize(220, 80)
        self.btn_next.clicked.connect(self.next_changes)
        self.btn_next.setEnabled(False)


        self.btn_add_part = QPushButton("Add Part")
        self.btn_add_part.setFixedSize(220, 80)
        self.btn_add_part.clicked.connect(self.open_add_part_dialog)
        self.btn_add_part.setEnabled(False)

        button_row1 = QHBoxLayout()
        button_row1.setAlignment(Qt.AlignmentFlag.AlignLeft)
        button_row1.addWidget(self.btn_open)
        button_row1.addWidget(self.btn_save)
        button_row1.addWidget(self.btn_next)
        button_row1.addWidget(self.btn_add_part)

        button_row2 = QHBoxLayout()
        button_row2.setAlignment(Qt.AlignmentFlag.AlignLeft)
        self.btn_reset = QPushButton("Reset Center Pose")
        self.btn_reset.setFixedSize(220, 80)
        self.btn_reset.clicked.connect(self.reset_center_pose)
        self.btn_reset.setEnabled(False)

        self.btn_set = QPushButton("Manual Set")
        self.btn_set.setFixedSize(220, 80)
        self.btn_set.clicked.connect(self.open_set_part_dialog)
        self.btn_set.setEnabled(False)
        button_row2.addWidget(self.btn_reset)
        button_row2.addWidget(self.btn_set)
        button_row2.addWidget(self.btn_open_excel)
        button_row2.addWidget(self.btn_open_excel_2) #yeni butonu ekle
        
        self.toggle_switch_name = QLabel("Excel Mode:")
        self.toggle_switch_name.setStyleSheet("""
            QLabel {
                color: #D8DEE9;
                font-size: 14px;
                margin-left: 5px;
            }
        """)
        # Toggle switch ve etiketi için layout
        toggle_container = QHBoxLayout()
        self.toggle_switch = ToggleSwitch()
        self.toggle_switch.setEnabled(False)  # Ana toggle switch her zaman aktif olacak
        self.toggle_label = QLabel("PART")
        self.toggle_label.setStyleSheet("""
            QLabel {
                color: #D8DEE9;
                font-size: 14px;
                margin-left: 5px;
            }
        """)
        
        toggle_container.addWidget(self.toggle_switch_name)
        toggle_container.addWidget(self.toggle_switch)
        toggle_container.addWidget(self.toggle_label)
        toggle_container.addStretch()
        
        # Toggle container'ı button_row2'ye ekle
        button_row2.addLayout(toggle_container)
        button_row1.addWidget(self.combo_excel_data)
        
        # Toggle switch sinyalini bağla
        self.toggle_switch.toggled.connect(self.on_toggle_switch)
        
        button_row1.addWidget(self.combo_excel_data)

        header_layout.addLayout(button_row1)
        header_layout.addLayout(button_row2)
        main_layout= QVBoxLayout()
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        self.container = QWidget()
        self.main_layout = QVBoxLayout()
        self.container.setLayout(self.main_layout)
        scroll.setWidget(self.container)
        scroll.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding)
        main_layout.addLayout(header_layout)
        main_layout.addWidget(scroll)
        self.setLayout(main_layout)

    def open_json_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Open JSON File", "", "JSON Files (*.json);;All Files (*)")
        if file_path:
            self.file_path = file_path
            self.lbl_file.setText(f"Loaded: {file_path}")
            try:
                with open(file_path, 'r', encoding='utf-8') as file:
                    self.json_data = json.load(file)
                    self.create_form()
                    self.btn_save.setEnabled(True)
                    self.btn_add_part.setEnabled(True)
                    self.btn_open_excel.setEnabled(True)
                    self.btn_set.setEnabled(True)
                    self.btn_reset.setEnabled(True)
                    self.btn_add_part.setEnabled(True)
                    self.btn_open_excel_2.setEnabled(True)
                    self.toggle_switch.setEnabled(False)
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Dosya açılamadı:\n{str(e)}")

    def load_excel_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select PART Excel File", "", "Excel Files (*.xlsx *.xls)")
        if not file_path:
            return
        self.excel_file_path_part = file_path
        try:
            self.excel_wb_part = openpyxl.load_workbook(file_path)
            sheet = self.excel_wb_part.active
            self.excel_data_part = []
            self.excel_headers_part = [sheet.cell(row=1, column=i).value for i in range(3, 6)]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(cell is not None for cell in row[2:5]):
                    self.excel_data_part.append(row[2:5])

            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                a_val = str(row[0]) if row[0] is not None else ""
                b_val = str(row[1]) if row[1] is not None else ""
                self.combo_excel_data.addItem(f"{a_val} - {b_val}")
        
            # Excel dosyası başarıyla yüklendiğinde toggle switch'i aktif et
            self.toggle_switch.setEnabled(True)
            QMessageBox.information(self, "Information", "PART Excel data loaded.")
        except Exception as e:
            self.toggle_switch.setEnabled(False)  # Hata durumunda toggle switch'i pasif et
            QMessageBox.critical(self, "Error", f"PART Excel file could not be uploaded:\n{str(e)}")

    def load_excel_file_plc(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select PLC Excel File", "", "Excel Files (*.xlsx *.xls)")
        if not file_path:
            return
        self.excel_file_path_plc = file_path
        try:
            self.excel_wb_plc = openpyxl.load_workbook(file_path)
            sheet = self.excel_wb_plc.active
            self.excel_data_plc = []
            self.excel_headers_plc = [sheet.cell(row=1, column=i).value for i in range(3, 6)]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(cell is not None for cell in row[2:5]):
                    self.excel_data_plc.append(row[2:5])

            # Excel dosyası başarıyla yüklendiğinde toggle switch'i aktif et
            self.toggle_switch.setEnabled(True)
            QMessageBox.information(self, "Information", "PLC Excel data loaded.")
        except Exception as e:
            self.toggle_switch.setEnabled(False)  # Hata durumunda toggle switch'i pasif et
            QMessageBox.critical(self, "Error", f"PLC Excel file could not be uploaded:\n{str(e)}")

    def on_excel_selection_changed(self,index):
        self.display_excel_row(index)
        

    def display_excel_row(self, index):
        if not self.auto_import_enabled:  # PART modunda ise
            if not hasattr(self, 'excel_data_part') or index < 0 or index >= len(self.excel_data_part):
                return

            # Önce eski ürün bilgilerini temizle
            for i in reversed(range(self.main_layout.count())):
                widget = self.main_layout.itemAt(i).widget()
                # Sadece daha önce eklenmiş ürün bilgisi kutularını kaldır
                if isinstance(widget, QGroupBox) and not widget.title().startswith("ROI"):
                    widget.setParent(None)

            # ComboBox'tan seçilen ürünün adını başlık olarak al
            selected_text = self.combo_excel_data.currentText()
            row_values = self.excel_data_part[index]
            group = QGroupBox(f"{selected_text}")  # Ürün adı başlık olarak
            layout = QFormLayout()

            headers = ["Uzunluk", "Genişlik", "Yükseklik"]
            for header, value in zip(headers, row_values):
                lbl = QLabel(header)
                line_edit = QLineEdit(str(value))
                line_edit.setReadOnly(True)
                layout.addRow(lbl, line_edit)

            group.setLayout(layout)
            self.main_layout.addWidget(group)
            self.btn_next.setEnabled(True)
        else:
            #PLC modunda hiçbir şey yapma
            pass

    def create_form(self):
        for i in reversed(range(self.main_layout.count())):
            widget = self.main_layout.itemAt(i).widget()
            if widget:
                widget.setParent(None)

        for section in self.input_fields.values():
            if isinstance(section, dict):
                for key in section:
                    section[key] = [] if isinstance(section[key], list) else None

        self.add_roi_section("roi3d_in_camera_coord", "ROI - Kamera Koordinatları")
        self.add_roi_section("roi3d_in_robot_coord", "ROI - Robot Koordinatları")

    def add_roi_section(self, key, title):
        group = QGroupBox(title)
        layout = QFormLayout()
        self.create_vector_inputs(key, "half_lengths", layout, self.json_data[key]["half_lengths"], "Half Lengths (X, Y, Z)")
        self.create_vector_inputs(key, "roi_center_pose", layout, self.json_data[key]["roi_center_pose"], "Center Pose (X,Y,Z,QX,QY,QZ,QW)")
        group.setLayout(layout)
        self.main_layout.addWidget(group)

    def create_vector_inputs(self, parent_key, child_key, layout, values, title):
        layout.addRow(QLabel(f"<b>{title}</b>"))
        components = ['X', 'Y', 'Z', 'QX', 'QY', 'QZ', 'QW']
        for i, value in enumerate(values):
            if i < len(components):
                lbl = QLabel(components[i])
                input_field = QLineEdit(self.format_float(value))
                input_field.setProperty("original_value", value)
                input_field.textChanged.connect(lambda text, field=input_field: self.on_text_changed(field))

                # + ve - butonlarını oluştur
                plus_button = QPushButton("+")
                minus_button = QPushButton("-")
                plus_button.setFixedSize(30, 25)
                minus_button.setFixedSize(30, 25)

                # Butonların tıklanma olaylarını bağla
                plus_button.clicked.connect(lambda _, field=input_field: self.adjust_value(field, 0.01))
                minus_button.clicked.connect(lambda _, field=input_field: self.adjust_value(field, -0.01))

                # Butonları ve input alanını yatay bir layout'a ekle
                hbox = QHBoxLayout()
                hbox.addWidget(input_field)
                hbox.addWidget(plus_button)
                hbox.addWidget(minus_button)

                layout.addRow(lbl, hbox)
                self.input_fields[parent_key][child_key].append(input_field)

    def adjust_value(self, field, percentage):
        try:
            current_value = float(field.text())
            adjustment = current_value * percentage
            new_value = current_value + adjustment
            field.setText(self.format_float(new_value))
        except ValueError:
            QMessageBox.warning(self, "Error", "Invalid numeric value, Please enter a valid number.")

    def on_text_changed(self, field):
        text = field.text()
        if self.is_invalid_float_format(text) or not self.is_valid_number(text):
            field.setStyleSheet("border: 2px solid #BF616A; background-color: #4C566A;")
            field.setProperty("error", "true")
        else:
            field.setStyleSheet("background-color: #3B4252; border: 1px solid #4C566A;")
            field.setProperty("error", "false")

    def format_float(self, value):
        num = float(value)
        return str(int(num)) if num.is_integer() else "{:.6f}".format(num).rstrip('0').rstrip('.')

    def is_valid_number(self, text):
        if text.strip() == "":
            return False
        try:
            float(text)
            if re.match(r'^\d+\.$', text) or re.match(r'^\.\d*$', text):
                return False
            return True
        except ValueError:
            return False

    def is_invalid_float_format(self, text):
        patterns = [
            r'^\d+\.$', r'^-\d+\.$', r'^\.\d*$', r'^-\.\d*$',
            r'^-?0+\d+$', r'^-?\.$', r'^[+-]?\d*\.0+$'
        ]
        return any(re.match(p, text) for p in patterns)

    def validate_all_fields(self):
        invalid = []
        for section_name, section in self.input_fields.items():
            if isinstance(section, dict):
                for field_name, fields in section.items():
                    if isinstance(fields, list):
                        for i, field in enumerate(fields):
                            if self.is_invalid_float_format(field.text()) or field.text().strip() == "":
                                field.setProperty("error", "true")
                                invalid.append(f"{section_name} > {field_name}")
                    elif isinstance(fields, QLineEdit):
                        if fields.text().strip() == "":
                            fields.setProperty("error", "true")
                            invalid.append(f"{section_name} > {field_name}")
        return invalid

    def save_changes(self):
        if self.validate_all_fields():
            QMessageBox.critical(self, "Error", "Please fill in all fields correctly.")
            return

        self.update_section("roi3d_in_camera_coord")
        self.update_section("roi3d_in_robot_coord")

        # OFFSET geçmişinin son değerini center pose Y'ye yaz
        if hasattr(self, "offset_history") and self.offset_history:
            self.json_data["roi3d_in_camera_coord"]["roi_center_pose"][1] = self.offset_history[-1]

        try:
            with open(self.file_path, 'w', encoding='utf-8') as f:
                json.dump(self.json_data, f, indent=4)
            QMessageBox.information(self, "Information", "Data saved successfully.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Cannot save :\n{str(e)}")

    def update_section(self, section_key):
        for key in ["half_lengths", "roi_center_pose"]:
            values = []
            for field in self.input_fields[section_key][key]:
                try:
                    val = float(field.text())
                    orig = field.property("original_value")
                    values.append(int(val) if isinstance(orig, int) and val.is_integer() else val)
                except ValueError:
                    values.append(field.property("original_value"))
            self.json_data[section_key][key] = values

    def open_add_part_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add new Part")
        layout = QFormLayout(dialog)

        part_code = QLineEdit()
        length = QLineEdit()
        width = QLineEdit()
        height = QLineEdit()

        layout.addRow("Part Code:", part_code)
        layout.addRow("Length:", length)
        layout.addRow("Width:", width)
        layout.addRow("Height:", height)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        layout.addWidget(buttons)

        def on_ok():
            code = part_code.text().strip()
            try:
                l = float(length.text().replace(",", "."))
                w = float(width.text().replace(",", "."))
                h = float(height.text().replace(",", "."))
            except ValueError:
                QMessageBox.warning(dialog, "Error", "Enter only numbers in the length, width and height fields!")
                return

            if not code:
                QMessageBox.warning(dialog, "Error", "The part code cannot be empty.")
                return

            # Excel dosyası yolu ve workbook kontrolü
            if not hasattr(self, "excel_file_path_part") or not self.excel_file_path_part:
                QMessageBox.warning(dialog, "Error", "First upload a PART Excel file.")
                return

            if not hasattr(self, "excel_wb_part") or self.excel_wb_part is None:
                QMessageBox.warning(dialog, "Error", "First upload a PART Excel file.")
                return

            sheet = self.excel_wb_part.active

            # 1. sütundaki en büyük sayıyı bul
            max_id = 0
            for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
                try:
                    val = int(row[0])
                    if val > max_id:
                        max_id = val
                except (TypeError, ValueError):
                    continue
            new_id = max_id + 1

            # Son satırı bul ve yeni satıra ekle
            last_row = sheet.max_row + 1
            sheet.cell(row=last_row, column=1).value = new_id
            sheet.cell(row=last_row, column=2).value = code
            sheet.cell(row=last_row, column=3).value = l
            sheet.cell(row=last_row, column=4).value = w
            sheet.cell(row=last_row, column=5).value = h

            try:
                self.excel_wb_part.save(self.excel_file_path_part)
            except PermissionError:
                QMessageBox.critical(dialog, "Error", "Cannot write to Excel file. Please make sure that the file is not open in another program and that you have write permission.")
                return

            # Hafızadaki dosya yolundan tekrar workbook'u yükle ve combobox'ı güncelle
            try:
                self.excel_wb_part = openpyxl.load_workbook(self.excel_file_path_part)
            except Exception as e:
                QMessageBox.critical(dialog, "Error", f"Excel could not be loaded again:\n{str(e)}")
                return

            sheet = self.excel_wb_part.active
            self.excel_data_part = []
            self.combo_excel_data.clear()
            self.excel_headers_part = [sheet.cell(row=1, column=i).value for i in range(3, 6)]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(cell is not None for cell in row[2:5]):
                    self.excel_data_part.append(row[2:5])
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                a_val = str(row[0]) if row[0] is not None else ""
                b_val = str(row[1]) if row[1] is not None else ""
                self.combo_excel_data.addItem(f"{a_val} - {b_val}")

            self.combo_excel_data.setCurrentIndex(self.combo_excel_data.count() - 1)
            QMessageBox.information(dialog, "Information", "Part added and Excel file updated.")
            dialog.accept()

        buttons.accepted.connect(on_ok)
        buttons.rejected.connect(dialog.reject)
        dialog.exec()

    def add_part_data(self, dialog, length_field, width_field, height_field):
        try:
            length = float(length_field.text())
            width = float(width_field.text())
            height = float(height_field.text())
        except ValueError:
            QMessageBox.warning(self, "Error", "All values must be numbers.")
            return

        new_part = {"length": length, "width": width, "height": height}
        if "extra_parts" not in self.json_data:
            self.json_data["extra_parts"] = []
        self.json_data["extra_parts"].append(new_part)

        try:
            with open(self.file_path, 'w', encoding='utf-8') as file:
                json.dump(self.json_data, file, indent=4)
            QMessageBox.information(self, "Information", "Part added and saved.")
            dialog.accept()
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kaydetme hatası:\n{str(e)}")

    def next_changes(self):
        # ROI kamera koordinatları center pose Y değerini al
        roi_center_pose = self.json_data.get("roi3d_in_camera_coord", {}).get("roi_center_pose", [])
        if len(roi_center_pose) < 2:
            QMessageBox.warning(self, "Warning", "ROI camera coordinates are missing.")
            return

        center_pose_y = self.json_data["roi3d_in_camera_coord"]["roi_center_pose"][1]

        index = self.combo_excel_data.currentIndex()
        if not hasattr(self, 'excel_data_part') or index < 0 or index >= len(self.excel_data_part):
            QMessageBox.warning(self, "Warning", "A valid part is not selected.")
            return

        part_width = self.excel_data_part[index][1]  # Genişlik sütunu

        try:
            # Genişlik bilgisini önce 1000'e böl, sonra çıkarma işlemi yap
            width_divided = float(part_width) / 1000
            offset = float(center_pose_y) - width_divided
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Calculation error:\n{str(e)}")
            return

        if not hasattr(self, "offset_history"):
            self.offset_history = []
        self.offset_history.append(offset)

        self.json_data["roi3d_in_camera_coord"]["roi_center_pose"][1] = offset

        # OFFSET hücresini ekrana ekle
        for i in reversed(range(self.main_layout.count())):
            widget = self.main_layout.itemAt(i).widget()
            if isinstance(widget, QGroupBox) and widget.title() == "OFFSET Results":
                widget.setParent(None)

        offset_group = QGroupBox("OFFSET Results")
        layout = QFormLayout()
        offset_edit = QLineEdit(str(offset))
        offset_edit.setReadOnly(True)
        layout.addRow(QLabel("OFFSET"), offset_edit)

        for idx, val in enumerate(self.offset_history, 1):
            edit = QLineEdit(str(val))
            edit.setReadOnly(True)
            layout.addRow(QLabel(f"OFFSET {idx}"), edit)

        offset_group.setLayout(layout)
        self.main_layout.addWidget(offset_group)

        self.json_data["offset_history"] = self.offset_history

        # Her ileri tuşunda otomatik kaydet
        try:
            with open(self.file_path, 'w', encoding='utf-8') as f:
                json.dump(self.json_data, f, indent=4)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Save Error:\n{str(e)}")

        # Formu güncelle, böylece Center Pose Y ekranda güncel görünür
        self.create_form()
        self.display_excel_row(index)

    def reset_center_pose(self):
        # Şifre penceresi aç
        dialog = QDialog(self)
        dialog.setWindowTitle("Password Required")
        layout = QFormLayout(dialog)

        password_edit = QLineEdit()
        password_edit.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addRow("Password:", password_edit)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        layout.addWidget(buttons)
            
        def on_ok():
            sifre = "muhtesemuclu"  
            if password_edit.text() == sifre:
                
                if "roi3d_in_camera_coord" in self.json_data and "roi_center_pose" in self.json_data["roi3d_in_camera_coord"]:
                    self.json_data["roi3d_in_camera_coord"]["roi_center_pose"] = [0, 0, 3, 0, 0, 0, 0]
                    try:
                        with open(self.file_path, 'w', encoding='utf-8') as f:
                            json.dump(self.json_data, f, indent=4)
                    except Exception as e:
                        QMessageBox.critical(self, "Error", f"Failed to save reset:\n{str(e)}")
                    self.create_form()
                    dialog.accept()  
                else:
                    QMessageBox.warning(self, "Warning", "ROI camera coordinates not found.")
            else:
                QMessageBox.warning(self, "Error", "Wrong Password!")

        buttons.accepted.connect(on_ok)
        buttons.rejected.connect(dialog.reject)
        dialog.exec()

    def open_set_part_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Set Center Pose")
        layout = QFormLayout(dialog)

        
        current_pose = self.json_data.get("roi3d_in_camera_coord", {}).get("roi_center_pose", [0,0,0,0,0,0,0])
        labels = ["X", "Y", "Z", "QX", "QY", "QZ", "QW"]
        edits = []

        for i, label in enumerate(labels):
            edit = QLineEdit(str(current_pose[i]) if i < len(current_pose) else "0")
            layout.addRow(f"{label}:", edit)
            edits.append(edit)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        layout.addWidget(buttons)

        def on_ok():
            try:
                new_pose = [float(edit.text()) for edit in edits]
                self.json_data["roi3d_in_camera_coord"]["roi_center_pose"] = new_pose
                with open(self.file_path, 'w', encoding='utf-8') as f:
                    json.dump(self.json_data, f, indent=4)
                self.create_form()  
                dialog.accept()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to update values:\n{str(e)}")

        buttons.accepted.connect(on_ok)
        buttons.rejected.connect(dialog.reject)

        dialog.exec()

    def on_toggle_switch(self, checked):
        if checked:
            # PLC aktif
            self.toggle_label.setText("PLC")
            
            self.auto_import_enabled = False
            self.auto_import_enabled_plc = True
            self.auto_import_timer.start(500)
            self.roi_update_timer.start(500)  # Timer'ı başlat
        else:
            # PART aktif
            self.toggle_label.setText("PART")
            self.auto_import_enabled = True
            self.auto_import_enabled_plc = False
            self.auto_import_timer.start(500)
            self.roi_update_timer.stop()  # Timer'ı durdur

    def auto_import_excel(self):
        
        if hasattr(self, "excel_file_path_part") and self.excel_file_path_part:
            try:
                self.excel_wb_part = openpyxl.load_workbook(self.excel_file_path_part)
                sheet = self.excel_wb_part.active
                self.excel_data_part = []
                self.excel_headers_part = [sheet.cell(row=1, column=i).value for i in range(3, 6)]
                exist_item=set()
                for i in range(self.combo_excel_data.count()):
                    exist_item.add(self.combo_excel_data.itemText(i))


                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if all(cell is not None for cell in row[2:5]):
                        self.excel_data_part.append(row[2:5])
                for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                    a_val = str(row[0]) if row[0] is not None else ""
                    b_val = str(row[1]) if row[1] is not None else ""
                    new_item=(f"{a_val} - {b_val}")
            
                if new_item not in exist_item:
                    self.combo_excel_data.addItem(new_item)
                    exist_item.add(new_item)
                self.display_excel_row(self.combo_excel_data.currentIndex())
            except Exception as e:
                pass  


    def auto_import_excel_plc(self):
        
         if hasattr(self, "excel_file_path_plc") and self.excel_file_path_plc:
            try:
                self.excel_wb_plc = openpyxl.load_workbook(self.excel_file_path_plc)
                sheet = self.excel_wb_plc.active
                self.excel_data_plc = []
                self.excel_headers_plc = [sheet.cell(row=1, column=i).value for i in range(3, 6)]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if all(cell is not None for cell in row[2:5]):
                        self.excel_data_plc.append(row[2:5])
                
                self.display_excel_row(self.combo_excel_data.currentIndex())
            except Exception as e:
                pass


    def update_roi_center_pose(self):
        if self.auto_import_enabled_plc and hasattr(self, "excel_file_path_plc") and self.excel_file_path_plc:
            try:
                # Her seferinde Excel dosyasını yeniden yükle
                try:
                    self.excel_wb_plc = openpyxl.load_workbook(self.excel_file_path_plc, read_only=False, data_only=True)
                except PermissionError:
                    print("The Excel file may be open by another program.")
                    return
                except Exception as e:
                    print(f"Error loading Excel file: {e}")
                    return

                if hasattr(self, "excel_wb_plc"):
                    sheet = self.excel_wb_plc.active
                    new_pose = []
                    for col in range(1, 8):
                        cell_value = sheet.cell(row=2, column=col).value
                        try:
                            new_pose.append(float(cell_value) if cell_value is not None else 0)
                        except ValueError:
                            new_pose.append(0)

                    # ROI kamera center pose değerlerini güncelle
                    self.json_data["roi3d_in_camera_coord"]["roi_center_pose"] = new_pose

                    # JSON dosyasını güncelle
                    try:
                        with open(self.file_path, 'w', encoding='utf-8') as f:
                            json.dump(self.json_data, f, indent=4)
                    except Exception as e:
                        QMessageBox.critical(self, "Error", f"Failed to update the JSON file:\n{str(e)}")
                        return

                    # Değişiklikleri ekrana yazdır
                    self.create_form()

                    # Excel dosyasını kapatmayı unutmayın
                    try:
                        self.excel_wb_plc.close()
                    except:
                        pass

            except Exception as e:
                print(f"General error:s {e}")
                try:
                    self.excel_wb_plc.close()
                except:
                    pass
    
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = RoiEditor()
    window.showMaximized()
    sys.exit(app.exec())